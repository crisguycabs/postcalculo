# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'mainProto2.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!

# pyuic5 -x mainUI.ui -o rei.py

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox
from tkinter import filedialog
import tkinter
import openpyxl
from pathlib import Path
import pandas as pd
from colParamChild import Ui_ColParamChild
from colQueryChild import Ui_ColQueryChild
from acercadeChild import Ui_Acercade
import json
import re

class Ui_MainProtoV2(object):
    def setupUi(self, MainProtoV2):
        #%% constructor
        MainProtoV2.setObjectName("MainProtoV2")
        MainProtoV2.resize(400, 355)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(MainProtoV2.sizePolicy().hasHeightForWidth())
        MainProtoV2.setSizePolicy(sizePolicy)
        MainProtoV2.setMinimumSize(QtCore.QSize(400, 355))
        MainProtoV2.setMaximumSize(QtCore.QSize(400, 355))
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("ecp.jpg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        MainProtoV2.setWindowIcon(icon)
        self.centralwidget = QtWidgets.QWidget(MainProtoV2)
        self.centralwidget.setObjectName("centralwidget")
        self.grpColumnas = QtWidgets.QGroupBox(self.centralwidget)
        self.grpColumnas.setEnabled(True)
        self.grpColumnas.setGeometry(QtCore.QRect(10, 120, 381, 191))
        self.grpColumnas.setObjectName("grpColumnas")
        self.lstColumnas = QtWidgets.QListWidget(self.grpColumnas)
        self.lstColumnas.setGeometry(QtCore.QRect(120, 20, 251, 131))
        self.lstColumnas.setObjectName("lstColumnas")
        self.btnEditar = QtWidgets.QPushButton(self.grpColumnas)
        self.btnEditar.setEnabled(True)
        self.btnEditar.setGeometry(QtCore.QRect(160, 160, 101, 23))
        self.btnEditar.setObjectName("btnEditar")
        self.btnBorrar = QtWidgets.QPushButton(self.grpColumnas)
        self.btnBorrar.setEnabled(True)
        self.btnBorrar.setGeometry(QtCore.QRect(270, 160, 101, 23))
        self.btnBorrar.setObjectName("btnBorrar")
        self.btnNuevoParametro = QtWidgets.QPushButton(self.grpColumnas)
        self.btnNuevoParametro.setEnabled(True)
        self.btnNuevoParametro.setGeometry(QtCore.QRect(10, 20, 101, 41))
        self.btnNuevoParametro.setObjectName("btnNuevoParametro")
        self.btnNuevoQuery = QtWidgets.QPushButton(self.grpColumnas)
        self.btnNuevoQuery.setEnabled(True)
        self.btnNuevoQuery.setGeometry(QtCore.QRect(10, 70, 101, 41))
        self.btnNuevoQuery.setObjectName("btnNuevoQuery")
        self.grpSabana = QtWidgets.QGroupBox(self.centralwidget)
        self.grpSabana.setEnabled(True)
        self.grpSabana.setGeometry(QtCore.QRect(10, 10, 381, 101))
        self.grpSabana.setObjectName("grpSabana")
        # self.label_4 = QtWidgets.QLabel(self.grpSabana)
        # self.label_4.setGeometry(QtCore.QRect(205, 45, 81, 16))
        # self.label_4.setObjectName("label_4")
        self.label_3 = QtWidgets.QLabel(self.grpSabana)
        self.label_3.setGeometry(QtCore.QRect(10, 20, 91, 16))
        self.label_3.setObjectName("label_3")
        self.cmbMetodo = QtWidgets.QComboBox(self.grpSabana)
        self.cmbMetodo.setEnabled(True)
        self.cmbMetodo.setGeometry(QtCore.QRect(255, 45, 113, 22))
        self.cmbMetodo.setObjectName("cmbMetodo")
        self.cmbMetodo.addItem("")
        self.cmbMetodo.addItem("")
        # self.txtCorrida = QtWidgets.QLineEdit(self.grpSabana)
        # self.txtCorrida.setEnabled(True)
        # self.txtCorrida.setGeometry(QtCore.QRect(255, 45, 113, 20))
        # self.txtCorrida.setObjectName("txtCorrida")
        self.label_5 = QtWidgets.QLabel(self.grpSabana)
        self.label_5.setGeometry(QtCore.QRect(205, 45, 81, 16))
        self.label_5.setObjectName("label_5")
        self.txtModelo = QtWidgets.QLineEdit(self.grpSabana)
        self.txtModelo.setEnabled(True)
        self.txtModelo.setGeometry(QtCore.QRect(60, 45, 113, 20))
        self.txtModelo.setObjectName("txtModelo")
        self.txtVersion = QtWidgets.QLineEdit(self.grpSabana)
        self.txtVersion.setEnabled(True)
        self.txtVersion.setGeometry(QtCore.QRect(60, 70, 113, 20))
        self.txtVersion.setObjectName("txtVersion")
        self.txtNombreSabana = QtWidgets.QLineEdit(self.grpSabana)
        self.txtNombreSabana.setEnabled(True)
        self.txtNombreSabana.setGeometry(QtCore.QRect(60, 20, 113, 20))
        self.txtNombreSabana.setReadOnly(True)
        self.txtNombreSabana.setObjectName("txtNombreSabana")
        self.label = QtWidgets.QLabel(self.grpSabana)
        self.label.setGeometry(QtCore.QRect(10, 45, 81, 16))
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.grpSabana)
        self.label_2.setGeometry(QtCore.QRect(10, 70, 91, 16))
        self.label_2.setObjectName("label_2")
        self.btnModAbstracto = QtWidgets.QPushButton(self.grpSabana)
        self.btnModAbstracto.setEnabled(False)
        self.btnModAbstracto.setGeometry(QtCore.QRect(255, 20, 113, 20))
        self.btnModAbstracto.setObjectName("btnModAbstracto")
        MainProtoV2.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainProtoV2)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 400, 21))
        self.menubar.setObjectName("menubar")
        self.menuArchivo = QtWidgets.QMenu(self.menubar)
        self.menuArchivo.setObjectName("menuArchivo")
        self.menuAcerca = QtWidgets.QMenu(self.menubar)
        self.menuAcerca.setObjectName("menuAcerca")
        MainProtoV2.setMenuBar(self.menubar)
        self.statusBar = QtWidgets.QStatusBar(MainProtoV2)
        self.statusBar.setObjectName("statusBar")
        MainProtoV2.setStatusBar(self.statusBar)
        self.actionNuevo_JSON = QtWidgets.QAction(MainProtoV2)
        self.actionNuevo_JSON.setObjectName("actionNuevo_JSON")
        self.actionAbrir_JSON = QtWidgets.QAction(MainProtoV2)
        self.actionAbrir_JSON.setObjectName("actionAbrir_JSON")
        self.actionGuardar_JSON = QtWidgets.QAction(MainProtoV2)
        self.actionGuardar_JSON.setEnabled(False)
        self.actionGuardar_JSON.setObjectName("actionGuardar_JSON")
        self.actionGuardar_como = QtWidgets.QAction(MainProtoV2)
        self.actionGuardar_como.setEnabled(False)
        self.actionGuardar_como.setObjectName("actionGuardar_como")
        self.actionCerrar = QtWidgets.QAction(MainProtoV2)
        self.actionCerrar.setObjectName("actionCerrar")
        self.actionAcerca_de = QtWidgets.QAction(MainProtoV2)
        self.actionAcerca_de.setObjectName("actionAcerca_de")
        self.actionAcerca_de_Modulo_de_Postcalculo = QtWidgets.QAction(MainProtoV2)
        self.actionAcerca_de_Modulo_de_Postcalculo.setObjectName("actionAcerca_de_Modulo_de_Postcalculo")
        self.menuArchivo.addAction(self.actionNuevo_JSON)
        self.menuArchivo.addAction(self.actionAbrir_JSON)
        self.menuArchivo.addSeparator()
        self.menuArchivo.addAction(self.actionGuardar_JSON)
        self.menuArchivo.addAction(self.actionGuardar_como)
        self.menuArchivo.addSeparator()
        self.menuArchivo.addAction(self.actionCerrar)
        self.menuAcerca.addAction(self.actionAcerca_de_Modulo_de_Postcalculo)
        self.menubar.addAction(self.menuArchivo.menuAction())
        self.menubar.addAction(self.menuAcerca.menuAction())

        self.retranslateUi(MainProtoV2)
        QtCore.QMetaObject.connectSlotsByName(MainProtoV2)
        
        #%% variables de clase y controles de la ventana
        
        self.grpColumnas.setEnabled(False)
        self.grpSabana.setEnabled(False)
        
        self.ruta_modelo=""
        self.ruta_Json=""
        self.nombre_sabana=""
        
        self.sabana={}
        
        self.sabana["MODELO"]=""
        self.sabana["VERSION"]=""
        # self.sabana["CORRIDA"]=""
        self.sabana["NOMBRE_SABANA"]=""
        self.sabana["DIMENSIONES"]=[]
        self.sabana["METODO"]=""
        
        self.data=[]
        
        self.ncol=0;
        self.nombre="COLUMNA"
        
        self.error_dialog = QtWidgets.QErrorMessage()
                
        self.dimensiones=list()
        
        #%% eventos de la ventana
        
        '''
        Crear un nuevo JSON
        '''
        self.actionNuevo_JSON.triggered.connect(self.NuevoJson)
        
        '''
        Crear una nueva columna tipo parametro
        '''
        self.btnNuevoParametro.clicked.connect(self.NuevaColumnaParam)
        
        '''
        Editar columna tipo parametro
        '''
        self.btnEditar.clicked.connect(self.EditarColumnaParam)
        
        '''
        crear una nueva columna tipo query
        '''
        self.btnNuevoQuery.clicked.connect(self.NuevaColumnaQuery)
        
        '''
        abrir acerca de
        '''
        self.actionAcerca_de_Modulo_de_Postcalculo.triggered.connect(self.Acercade)
        
        '''
        cerrar la aplicación
        '''
        # self.actionCerrar.triggered.connect(self.closewin)
        
        '''
        guardar json
        '''
        self.actionGuardar_JSON.triggered.connect(self.GuardarJSON)
        
        '''
        guardar json como
        '''
        self.actionGuardar_como.triggered.connect(self.GuardarJSONcomo)
        
        '''
        abrir json
        '''
        self.actionAbrir_JSON.triggered.connect(self.AbrirJSON)
        
        '''
        seleccionar el modelo abstracto
        '''
        self.btnModAbstracto.clicked.connect(self.SelModelo)
        
        '''
        borrar columna
        '''
        self.btnBorrar.clicked.connect(self.EliminarColumna)
        
        #%% retranslateUI

    def retranslateUi(self, MainProtoV2):
        _translate = QtCore.QCoreApplication.translate
        MainProtoV2.setWindowTitle(_translate("MainProtoV2", "Modulo de PostCalculo BetaV2"))
        self.grpColumnas.setTitle(_translate("MainProtoV2", "Colección de Columnas"))
        self.btnEditar.setText(_translate("MainProtoV2", "Editar Columna"))
        self.btnBorrar.setText(_translate("MainProtoV2", "Eliminar Columna"))
        self.btnNuevoParametro.setText(_translate("MainProtoV2", "Nueva columna \n"
"parametro\\formula"))
        self.btnNuevoQuery.setText(_translate("MainProtoV2", "Nueva columna \n"
"query"))
        self.grpSabana.setTitle(_translate("MainProtoV2", "Sabana"))
        # self.label_4.setText(_translate("MainProtoV2", "Corrida: "))
        self.label_3.setText(_translate("MainProtoV2", "Nombre: "))
        self.cmbMetodo.setItemText(0, _translate("MainProtoV2", "MERGE"))
        self.cmbMetodo.setItemText(1, _translate("MainProtoV2", "COMBINAR"))
        # self.txtCorrida.setPlaceholderText(_translate("MainProtoV2", "Por asignar..."))
        self.label_5.setText(_translate("MainProtoV2", "Metodo: "))
        self.txtModelo.setPlaceholderText(_translate("MainProtoV2", "Por asignar..."))
        self.txtVersion.setPlaceholderText(_translate("MainProtoV2", "Por asignar..."))
        self.txtNombreSabana.setPlaceholderText(_translate("MainProtoV2", "Por asignar..."))
        self.label.setText(_translate("MainProtoV2", "Modelo:"))
        self.label_2.setText(_translate("MainProtoV2", "Version:"))
        self.btnModAbstracto.setText(_translate("MainProtoV2", "Seleccionar Modelo"))
        self.menuArchivo.setTitle(_translate("MainProtoV2", "Archivo"))
        self.menuAcerca.setTitle(_translate("MainProtoV2", "Acerca de..."))
        self.actionNuevo_JSON.setText(_translate("MainProtoV2", "Nuevo JSON"))
        self.actionAbrir_JSON.setText(_translate("MainProtoV2", "Abrir JSON"))
        self.actionGuardar_JSON.setText(_translate("MainProtoV2", "Guardar JSON"))
        self.actionGuardar_como.setText(_translate("MainProtoV2", "Guardar como..."))
        self.actionCerrar.setText(_translate("MainProtoV2", "Cerrar"))
        self.actionAcerca_de.setText(_translate("MainProtoV2", "Acerca de..."))
        self.actionAcerca_de_Modulo_de_Postcalculo.setText(_translate("MainProtoV2", "Acerca de Modulo de Postcalculo"))


    #%% metodos custom
    
    '''
    Eliminar columnas
    '''
    def EliminarColumna(self):
        
        msgbox = QMessageBox(QMessageBox.Question, "Confirmacion", "Desea eliminar la columna seleccionada?")
        msgbox.addButton(QMessageBox.Yes)
        msgbox.addButton(QMessageBox.No)
        msgbox.setDefaultButton(QMessageBox.No)
        reply = msgbox.exec()
        if reply != QMessageBox.Yes:
                return
        
        # no hay elementos a editar
        if(self.lstColumnas.count()<1):
            self.statusBar.showMessage("No existe columna a eliminar...")
            return
        
        nombre="COLUMNA" + str(self.lstColumnas.currentRow()+1)
        self.sabana.pop(nombre)
        self.ncol=self.ncol-1
        
        # se obtienen la lista de llaves "COLUMNA" restantes
        llaves=list(self.sabana.keys())
        subllaves=[x for x in llaves if re.search("COLUMNA", x)]    
        
        # se renombran las llaves para mantener la secuencia de 1 en adelante
        for i in range(len(subllaves)):
            vieja=subllaves[i]
            nueva="COLUMNA" + str(i+1)
            if(vieja!=nueva):
                self.sabana[nueva]=self.sabana[vieja]
                del self.sabana[vieja]
        
        self.PrintColumnas()
    
    '''
    Editar columna parametro/formula
    '''
    def EditarColumnaParam(self):
        
        # no hay elementos a editar
        if(self.lstColumnas.count()<1):
            self.statusBar.showMessage("No existe columna a editar...")
            return
        
        # no hay modelo seleccionado
        if(self.ruta_modelo==""):
            self.statusBar.showMessage("No se pueden crear/modificar columnas hasta que no se escoja un modelo abstracto")
            self.error_dialog.showMessage("No se pueden crear/modificar columnas hasta que no se escoja un modelo abstracto")
            return
        
        ncurrent="COLUMNA" + str(self.lstColumnas.currentRow()+1)
        current=self.sabana[ncurrent][0]
        
        if(current["TIPO_DATO_GAMS"]=="QUERY"):
            # se abre la ventana de columna query
            myColQueryChild=QtWidgets.QDialog()
            self.uiQuery=Ui_ColQueryChild()
            self.uiQuery.setupUi(myColQueryChild)
            
            self.uiQuery.SetCustom(self)
            
            myColQueryChild.setWindowTitle("Editar columna query")
            myColQueryChild.setModal(True)
            
            if(self.dimensiones):
                if(type(self.dimensiones)==list):
                    self.uiQuery.lblDimensiones.setText("Existe un conjunto previo de dimensiones: " + ",".join(self.dimensiones))
                else:
                    self.uiQuery.lblDimensiones.setText("Existe un conjunto previo de dimensiones: " + self.dimensiones)
            else:
                self.uiQuery.lblDimensiones.setText("")
                
            self.uiQuery.txtNombre.setText(current["NOMBRE_COLUMNA"])
            self.uiQuery.txtSql.setText(current["FORMULA"])
            
            self.uiQuery.txtDimensiones.setText(",".join(self.dimensiones))
            
            if(myColQueryChild.exec()==QtWidgets.QDialog.Accepted):
                self.sabana[ncurrent][0]["NOMBRE_COLUMNA"]=self.uiQuery.txtNombre.text()
                self.sabana[ncurrent][0]["FORMULA"]=self.uiQuery.txtSql.text()
                self.PrintColumnas()
            else:
                self.statusBar.showMessage("Se canceló la edición de la columna")
        else:
            # se abre la ventana de columna parametro/formula
            myColParamChild = QtWidgets.QDialog()
            self.uiParam = Ui_ColParamChild()
            self.uiParam.setupUi(myColParamChild)
            
            self.uiParam.SetCustom(self)
            
            myColParamChild.setWindowTitle("Editar columna parametro/formula")                
            myColParamChild.setModal(True)
            
            if(current["TIPO_DATO_GAMS"]=="FORMULA"):
                self.uiParam.chkFormula.setChecked(True)
                self.uiParam.txtFormula.setText(self.sabana[ncurrent][0]["FORMULA"])
            else:
                self.uiParam.chkFormula.setChecked(False)
            
            self.uiParam.txtNombreColumna.setText(self.sabana[ncurrent][0]["NOMBRE_COLUMNA"])
            
            elementos=self.sabana[ncurrent][0]["NOMBRE_E_S"]
            
            for i in range(self.uiParam.tableWidget.rowCount()):
                item = self.uiParam.tableWidget.item(i, 1)
                for j in range(len(elementos)):
                    if(item.text()==elementos[j]):
                        self.uiParam.tableWidget.item(i, 0).setCheckState(QtCore.Qt.Checked)
                    
                    
            if(myColParamChild.exec()==QtWidgets.QDialog.Accepted):
                
                self.statusBar.showMessage("Editando columna parámetro...")
                
                # verificaciones de datos ingresados
                
                # parametro o formula
                if(self.uiParam.chkFormula.isChecked()):
                    _tipoDato="FORMULA"
                    _formula=self.uiParam.txtFormula.toPlainText()
                else:
                    _tipoDato="PARAMETER"
                    _formula="NA"
                
                # nombre de columna
                if(self.uiParam.txtNombreColumna.text()):
                    _nombreCol=self.uiParam.txtNombreColumna.text()
                else:
                    self.statusBar.showMessage("No se incluyó un nombre de la columna")
                    self.error_dialog.showMessage("No se incluyó un nombre de la columna. Se cancela la creación de la nueva columna")
                    return
                
                # parametros seleccionados
                lista=list()
                for i in range(self.uiParam.tableWidget.rowCount()):
                    item = self.uiParam.tableWidget.item(i, 0)
                    if item.checkState() == QtCore.Qt.Checked:
                        # es un parametro o variable?
                        temp=self.data.loc[self.data["nombre_dato"]==self.uiParam.tableWidget.item(i, 1).text()]
                        lista.append((self.uiParam.tableWidget.item(i, 1).text(),self.uiParam.tableWidget.item(i, 2).text(),temp.iloc[0,0]))
                print(lista)
                
                if(len(lista)<1):
                    self.statusBar.showMessage("No se escogió ningún parámetro. Modifique su selección")
                    self.error_dialog.showMessage("No se escogió ningún parámetro. Modifique su selección")
                    return                
                
                _tipoCampo=list()
                _nombre=list()
                for i in range(len(lista)):
                    if(lista[i][2]=="parameter"):
                        _tipoCampo.append("value")
                    else:
                        _tipoCampo.append("level")
                        if not self.uiParam.chkFormula.isChecked():
                            _tipoDato="VARIABLE"
                    _nombre.append(lista[i][0])
                    
                # dimensiones
                if(len(lista)>=1):
                    self.dimensiones=lista[0][1]
                    self.sabana["DIMENSIONES"]=self.dimensiones
                    if(len(lista)>1):
                        for i in range(1,len(lista)):
                            for j in range(0,len(lista)-1):
                                if(lista[i][1]!=lista[j][1]):
                                    self.statusBar.showMessage("Las dimensiones seleccionadas no coinciden. Modifique su selección")
                                    self.error_dialog.showMessage("Las dimensiones seleccionadas no coinciden. Modifique su selección")
                                    return
                
                self.sabana[ncurrent][0]["TIPO_DATO_GAMS"]=_tipoDato
                self.sabana[ncurrent][0]["NOMBRE_E_S"]=_nombre
                self.sabana[ncurrent][0]["TIPO_CAMPO"]=_tipoCampo
                self.sabana[ncurrent][0]["NOMBRE_COLUMNA"]=_nombreCol
                self.sabana[ncurrent][0]["FORMULA"]=_formula
                
                self.PrintColumnas()
            else:
                self.statusBar.showMessage("Se cancela la edición...")
            
    
    '''
    Seleccionar modelo abstracto
    '''
    def SelModelo(self):
        root = tkinter.Tk()
        root.withdraw()
        
        self.statusBar.showMessage("Seleccionando modelo abstracto...")
        
        file_path1 = filedialog.askopenfilename(title="Seleccionando modelo abstracto...",filetypes=[('Excel', '*.xlsx'),('Excel', '*.xls')])
        if file_path1 is None:
            return None    
        
        self.ruta_modelo=file_path1
        
        # extraer la data   
        self.statusBar.showMessage("Procesando el modelo abstracto...")
        
        xlsx_file = Path('', self.ruta_modelo)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active
        
        self.l=[[0 for j in range(4)] for i in range(sheet.max_row)]
        self.l[0]=("tipo_dato","nombre_dato","dimensiones","categoria")                
        
        n=1
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i != 0:
                word=row[1]
                if (word.find("VARIABLE")!=-1):
                    
                    # if(row[6]==None):
                    #     row[6]="NA"
                    # if(row[3]==None):
                    #     row[3]="NA"
                        
                    self.l[n]=("variable",row[2],row[3],row[6])                    
                    n=n+1
                    
                elif ((word.find("PARAMETER")!=-1) or (word.find("TABLE")!=-1)):
                    
                    # if(row[6]==None):
                    #     row[6]="NA"
                    # if(row[3]==None):
                    #     row[3]="NA"
                    
                    self.l[n]=("parameter",row[2],row[3],row[6])
                    n=n+1    

        # algunas filas quedan vacias, se eliminan
        self.l[n:sheet.max_row]=[]
                
        self.data=pd.DataFrame(data=self.l[1:], index=None, columns=self.l[0])
        
        # se corrigen los NoneType
        for i in range(len(self.data)):
            if self.data.iloc[i,2] is None:
                self.data.iloc[i,2]=""
            if self.data.iloc[i,3] is None:
                self.data.iloc[i,3]=""
        
        self.statusBar.showMessage("Listo!")
        
        self.btnModAbstracto.setEnabled(False)
        
    '''
    Abrir JSON
    '''
    def AbrirJSON(self):
        root = tkinter.Tk()
        root.withdraw()
        
        self.statusBar.showMessage("Seleccionando archivo JSON...")
        
        # escoger la ruta de guardado del JSON
        # file_path1 = filedialog.asksaveasfile(title="Abriendo JSON",mode='r', defaultextension=".json")
        file_path1 = filedialog.askopenfilename(title="Abriendo JSON",filetypes=[('JSON', '*.json')])
        if file_path1 is None:
            return None        
        
        if len(file_path1)==0:
            self.statusBar.showMessage("Se canceló la lectura del JSON")
            return None
        
        self.ruta_Json=file_path1
        
        try:
            with open(self.ruta_Json) as json_file:
                self.sabana = json.load(json_file)
            self.statusBar.showMessage("Leyendo JSON...")
        except:
            self.statusBar.showMessage("Error al intentar leer JSON!")
            return None
        
        self.LimpiarControles()
        
        self.nombre_sabana=self.sabana["NOMBRE_SABANA"]
        
        self.txtNombreSabana.setText(self.sabana["NOMBRE_SABANA"])
        self.txtModelo.setText(self.sabana["MODELO"])
        self.txtVersion.setText(self.sabana["VERSION"])
        # self.txtCorrida.setText(self.sabana["CORRIDA"])
        if(self.sabana["METODO"]=="MERGE"):
            self.cmbMetodo.setCurrentIndex(0)
        else:
            self.cmbMetodo.setCurrentIndex(1)
            
        self.ncol=0
        sw=True
        while(sw):
            nombre="COLUMNA" + str(self.ncol+1)
            if nombre in self.sabana:
                self.ncol=self.ncol+1
            else:
                sw=False
                
        self.dimensiones=self.sabana["DIMENSIONES"]
        self.dimensiones=",".join(self.dimensiones)
        
        self.PrintColumnas()    
        
        self.btnModAbstracto.setEnabled(True)
        
        self.ruta_modelo=""
        
        self.statusBar.showMessage("JSON cargado con exito!")
        
        self.actionGuardar_JSON.setEnabled(True)
        self.actionGuardar_como.setEnabled(True)
    
        
    '''
    Guardar JSON en disco como
    '''
    def GuardarJSONcomo(self):
        root = tkinter.Tk()
        root.withdraw()
        
        self.statusBar.showMessage("Seleccionando archivo JSON...")
        
        # escoger la ruta de guardado del JSON
        file_path1 = filedialog.asksaveasfile(title="Guardar nuevo JSON",mode='w', defaultextension=".json")
        if file_path1 is None:
            return None        
        
        file_path1.close()
        
        self.ruta_Json=file_path1.name
        
        self.nombre_sabana=self.ruta_Json.split("/")
        self.nombre_sabana=self.nombre_sabana[-1]
        self.nombre_sabana=self.nombre_sabana.split(".")
        self.nombre_sabana=self.nombre_sabana[0]
        
        self.txtNombreSabana.setText(self.nombre_sabana)
        
        self.GuardarJSON()
    
    '''
    Guardar JSON en disco
    '''
    def GuardarJSON(self):
        
        # verificacion de elementos de entrada
        
        # el nombre del json ya se ha definido en el momento de creacion
        
        # nombre del modelo
        if(self.txtModelo.text()):
            self.sabana["MODELO"]=self.txtModelo.text()
        else:
            self.statusBar.showMessage("No se incluyó un nombre de modelo")
            self.error_dialog.showMessage("No se incluyó un nombre de modelo. Se cancela la operación de guardado")
            return
        
        # nombre de la version
        if(self.txtVersion.text()):
            self.sabana["VERSION"]=self.txtVersion.text()
        else:
            self.statusBar.showMessage("No se incluyó un nombre de version")
            self.error_dialog.showMessage("No se incluyó un nombre de version. Se cancela la operación de guardado")
            return
        
        # # nombre de la corrida
        # if(self.txtCorrida.text()):
        #     self.sabana["CORRIDA"]=self.txtCorrida.text()
        # else:
        #     self.statusBar.showMessage("No se incluyó un nombre de corrida")
        #     self.error_dialog.showMessage("No se incluyó un nombre de corrida. Se cancela la operación de guardado")
        #     return
        
        # columnas 
        if(self.lstColumnas.count()<1):
            self.statusBar.showMessage("No se han incluido columnas")
            self.error_dialog.showMessage("No se han incluido columnas. Se cancela la operación de guardado")
            return
        
        # nombre sabana
        self.sabana["NOMBRE_SABANA"]=self.nombre_sabana
        
        # metodo
        self.sabana["METODO"]=self.cmbMetodo.currentText()
        
        # dimensiones
        if(type(self.dimensiones)!=list):
            tdim=self.dimensiones.split(",")
        else:
            tdim=self.dimensiones
        self.sabana["DIMENSIONES"]=tdim
        
        sabanaJSON=json.dumps(self.sabana,indent=4)
        with open(self.ruta_Json, 'w') as outfile:
            outfile.write(sabanaJSON)
            
        self.statusBar.showMessage("JSON guardado con exito")
    
    '''
    Se limpian todos los controles, reseteo de la pantalla
    '''
    def LimpiarControles(self):
        self.grpColumnas.setEnabled(True)
        self.grpSabana.setEnabled(True)
        # self.txtCorrida.setText("")
        self.txtModelo.setText("")
        self.txtNombreSabana.setText("")
        self.txtVersion.setText("")
        self.lstColumnas.clear()
        self.cmbMetodo.setCurrentIndex(0)        

    '''
    Se crea un nuevo JSON, desde cero. Se pide la ruta del modelo abstracto y 
    la ruta donde se deberia guardar el JSON
    '''
    def NuevoJson(self):
        
        root = tkinter.Tk()
        root.withdraw()
        
        self.statusBar.showMessage("Seleccionando archivo JSON...")
        
        # escoger la ruta de guardado del JSON
        file_path1 = filedialog.asksaveasfile(title="Guardar nuevo JSON",mode='w', defaultextension=".json")
        if (file_path1 is None):
            return None        
        
        file_path1.close()
        
        self.statusBar.showMessage("Seleccionando modelo abstracto...")
        
        # escoger el modelo abstracto
        file_path2 = filedialog.askopenfilename(title='Seleccione el modelo abstracto...',
                                                        filetypes=[("Archivos excel", ".xlsx .xls")])
        
        
        if (file_path2 is None) or (len(file_path2) == 0):
            self.statusBar.showMessage("Se canceló la creación del nuevo JSON.")
            return None
        
        # limpiando variables y cache
        self.ruta_Json=""
        self.ruta_modelo=""
        self.nombre_sabana=""
        self.data=""
        self.sabana=""
        self.ncol=0
        self.dimensiones=""
        self.nombre=""
        self.l=0
        
        self.statusBar.showMessage("Procesando modelo abstracto...")
        
        self.ruta_Json=file_path1.name
        self.ruta_modelo=file_path2
        
        self.nombre_sabana=self.ruta_Json.split("/")
        self.nombre_sabana=self.nombre_sabana[-1]
        self.nombre_sabana=self.nombre_sabana.split(".")
        self.nombre_sabana=self.nombre_sabana[0]
        
        self.LimpiarControles()
        
        self.actionGuardar_JSON.setEnabled(True)
        self.actionGuardar_como.setEnabled(True)
        
        self.txtNombreSabana.setText(self.nombre_sabana)
        
        # extraer la data   
        xlsx_file = Path('', self.ruta_modelo)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active
        
        self.l=[[0 for j in range(4)] for i in range(sheet.max_row)]
        self.l[0]=("tipo_dato","nombre_dato","dimensiones","categoria")                
        
        n=1
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i != 0:
                word=row[1]
                sw=False
                if (word.find("VARIABLE")!=-1):
                    
                    tipo_dato="variable"
                    sw=True                    
                    
                elif ((word.find("PARAMETER")!=-1) or (word.find("TABLE")!=-1)):
                    
                    tipo_dato="parameter"
                    sw=True
                    
                if sw==True:   
                    nombre_dato = row[2]
                    if nombre_dato==None:
                        nombre_dato=""
                        
                    dimensiones = row[3]
                    if dimensiones==None:
                        dimensiones=""
                        
                    categoria = row[6]
                    if categoria==None:
                        categoria="Sin Categoria"
                        
                    self.l[n]=(tipo_dato,nombre_dato,dimensiones,categoria)
                        
                    n=n+1                
                

        # algunas filas quedan vacias, se eliminan
        self.l[n:sheet.max_row]=[]
                
        self.data=pd.DataFrame(data=self.l[1:], index=None, columns=self.l[0])
        
        # # se corrigen los NoneType
        # for i in range(len(self.data)):
        #     if self.data.iloc[i,2] is None:
        #         self.data.iloc[i,2]=""
        #     if self.data.iloc[i,3] is None:
        #         self.data.iloc[i,3]=""
        
        self.statusBar.showMessage("Listo!")
    
    '''
    ventana acerca de
    '''
    def Acercade(self):
        Acercade = QtWidgets.QDialog()
        ui = Ui_Acercade()
        ui.setupUi(Acercade)
        Acercade.setModal(True)
        
        if(Acercade.exec()==QtWidgets.QDialog.Accepted):
            print("Dialogo nueva columna cerrado con exito")
            
    '''
    Cerrar la aplicación
    '''        
    def exit_window(self, event):
        close = QtWidgets.QMessageBox.question(self,
                                            "Salir?",
                                            "Está seguro que desea salir?",
                                            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        if close == QtWidgets.QMessageBox.Yes:
            # event.accept()
            sys.exit()
        else:
            pass
        
    '''
    Se abre la ventana de nueva columna tip query
    '''
    def NuevaColumnaQuery(self):
        
        if(self.ruta_modelo==""):
            self.statusBar.showMessage("No se pueden crear/modificar columnas hasta que no se escoja un modelo abstracto")
            self.error_dialog.showMessage("No se pueden crear/modificar columnas hasta que no se escoja un modelo abstracto")
            return
        
        myColQueryChild=QtWidgets.QDialog()
        self.uiQuery=Ui_ColQueryChild()
        self.uiQuery.setupUi(myColQueryChild)
        
        self.uiQuery.SetCustom(self)
        
        myColQueryChild.setWindowTitle("Nueva columna query")
        myColQueryChild.setModal(True)
        
        if(self.dimensiones):
            self.uiQuery.lblDimensiones.setText("Existe un conjunto previo de dimensiones: " + self.dimensiones)
        else:
            self.uiQuery.lblDimensiones.setText("")
        
        if(myColQueryChild.exec()==QtWidgets.QDialog.Accepted):
            
            self.statusBar.showMessage("Procesando nueva columna parámetro...")
            
            temCol=[]
            columna={}
            
            # verificaciones de datos ingresados
            
            # nombre de columna
            if(self.uiQuery.txtNombre.text()):
                _nombreCol=self.uiQuery.txtNombre.text()
            else:
                self.statusBar.showMessage("No se incluyó un nombre de la columna")
                self.error_dialog.showMessage("No se incluyó un nombre de la columna. Se cancela la creación de la nueva columna")
                return
            
            # dimensiones
            if(self.uiQuery.txtDimensiones.text()):
                if(self.dimensiones):
                    # hay dimensiones previas
                    if(self.dimensiones!=self.uiQuery.txtDimensiones.text()):
                        self.statusBar.showMessage("Las dimensiones incluidas no coinciden con las previas")
                        self.error_dialog.showMessage("Las dimensiones incluidas no coinciden con las previas. Se cancela la creación de la nueva columna")
                        return
                else:
                    # no hay dimensiones previas
                    self.dimensiones=self.uiQuery.txtDimensiones.text()
            else:
                self.statusBar.showMessage("No se incluyeron dimensiones en la columna")
                self.error_dialog.showMessage("No se incluyeron dimensiones en la columna. Se cancela la creación de la nueva columna")
                return
            
            # sentencia SQL
            if(self.uiQuery.txtSql.toPlainText()):
                _formula=self.uiQuery.txtSql.toPlainText()
            else:
                self.statusBar.showMessage("No se incluyó una instrucción SQL")
                self.error_dialog.showMessage("No se incluyó una instrucción SQL. Se cancela la creación de la nueva columna")
                return
            
            columna["TIPO_DATO_GAMS"]="QUERY"
            columna["NOMBRE_E_S"]=[]
            columna["TIPO_CAMPO"]=[]
            columna["NOMBRE_COLUMNA"]=_nombreCol
            columna["FORMULA"]=_formula
            temCol.append(columna)
            
            self.ncol=self.ncol+1
            col=self.nombre+str(self.ncol)
            self.sabana[col]=temCol
            
            self.PrintColumnas()
            
            self.statusBar.showMessage("Nueva columna creada!")
    '''
    Se recorren todas las columnas y se agregan al control listbox
    '''
    def PrintColumnas(self):
        
        self.lstColumnas.clear()
        
        for i in range(1,self.ncol+1):
            nomcol=self.nombre+str(i)
                        
            # se agrega la columna a la lista
            item=""
            # item=item + "Nombre: " + self.sabana[nomcol][0]["NOMBRE_COLUMNA"] + " | "
            # item=item + "Tipo: " + self.sabana[nomcol][0]["TIPO_DATO_GAMS"]
            item = item + self.sabana[nomcol][0]["NOMBRE_COLUMNA"] + " | "
            if isinstance(self.sabana[nomcol][0]["NOMBRE_E_S"], list):
                for j in range(len(self.sabana[nomcol][0]["NOMBRE_E_S"])):
                    item = item + self.sabana[nomcol][0]["NOMBRE_E_S"][j] + " "
                item = item + " | "
            else:
                item = item + self.sabana[nomcol][0]["NOMBRE_E_S"] + " | "
            item = item + self.sabana[nomcol][0]["TIPO_DATO_GAMS"]
            self.lstColumnas.addItem(item)
       
    
    '''
    Se abre la ventana de nueva columna tipo parametro/formula
    '''
    def NuevaColumnaParam(self):
        
        if(self.ruta_modelo==""):
            self.statusBar.showMessage("No se pueden crear/modificar columnas hasta que no se escoja un modelo abstracto")
            self.error_dialog.showMessage("No se pueden crear/modificar columnas hasta que no se escoja un modelo abstracto")
            return
        
        myColParamChild = QtWidgets.QDialog()
        self.uiParam = Ui_ColParamChild()
        self.uiParam.setupUi(myColParamChild)
        
        self.uiParam.SetCustom(self)
        
        myColParamChild.setWindowTitle("Nueva columna parametro/formula")                
        myColParamChild.setModal(True)
        
        if(myColParamChild.exec()==QtWidgets.QDialog.Accepted):
            
            self.statusBar.showMessage("Procesando nueva columna parámetro...")
            
            temCol=[]
            columna={}
            
            # verificaciones de datos ingresados
            
            # parametro o formula
            if(self.uiParam.chkFormula.isChecked()):
                _tipoDato="FORMULA"
                _formula=self.uiParam.txtFormula.toPlainText()
            else:
                _tipoDato="PARAMETER"
                _formula="NA"
            
            # nombre de columna
            if(self.uiParam.txtNombreColumna.text()):
                
                # se verifica que no exista que no exista una columna con un nombre similar
                for i in range(self.ncol):
                    tempnom = "COLUMNA" + str(i+1)
                    if (self.sabana[tempnom][0]["NOMBRE_COLUMNA"]==self.uiParam.txtNombreColumna.text()):
                        self.statusBar.showMessage("Nombre de columna inválido...")
                        self.error_dialog.showMessage("Ya existe una columna con el nombre " + self.uiParam.txtNombreColumna.text() + ". Se cancela la creación de la nueva columna")
                        return None                             
                
                _nombreCol=self.uiParam.txtNombreColumna.text()
            else:
                
                # se verifica que solo exista un unico parametro seleccionado para darle el nombre a la columna
                lista=list()
                for i in range(self.uiParam.tableWidget.rowCount()):
                    item = self.uiParam.tableWidget.item(i, 0)
                    if item.checkState() == QtCore.Qt.Checked:
                        # es un parametro o variable?
                        lista.append(self.uiParam.tableWidget.item(i, 1).text())
                
                if len(lista)==1:
                    _nombreCol = lista[0]
                else:                
                    self.statusBar.showMessage("No se incluyó un nombre de la columna")
                    self.error_dialog.showMessage("No se incluyó un nombre de la columna. Se cancela la creación de la nueva columna")
                    return
            
            # parametros seleccionados
            lista=list()
            for i in range(self.uiParam.tableWidget.rowCount()):
                item = self.uiParam.tableWidget.item(i, 0)
                if item.checkState() == QtCore.Qt.Checked:
                    # es un parametro o variable?
                    temp=self.data.loc[self.data["nombre_dato"]==self.uiParam.tableWidget.item(i, 1).text()]
                    lista.append((self.uiParam.tableWidget.item(i, 1).text(),self.uiParam.tableWidget.item(i, 2).text(),temp.iloc[0,0]))
            print(lista)
            
            if(len(lista)<1):
                self.statusBar.showMessage("No se escogió ningún parámetro. Modifique su selección")
                self.error_dialog.showMessage("No se escogió ningún parámetro. Modifique su selección")
                return                
            
            _tipoCampo=list()
            _nombre=list()
            for i in range(len(lista)):
                if(lista[i][2]=="parameter"):
                    _tipoCampo.append("value")
                else:
                    _tipoCampo.append("level")
                    if not self.uiParam.chkFormula.isChecked():
                        _tipoDato="VARIABLE"
                _nombre.append(lista[i][0])
                
            # dimensiones
            if(len(lista)>=1):
                self.dimensiones=lista[0][1]
                self.sabana["DIMENSIONES"]=self.dimensiones
                if(len(lista)>1):
                    for i in range(1,len(lista)):
                        for j in range(0,len(lista)-1):
                            if(lista[i][1]!=lista[j][1]):
                                self.statusBar.showMessage("Las dimensiones seleccionadas no coinciden. Modifique su selección")
                                self.error_dialog.showMessage("Las dimensiones seleccionadas no coinciden. Modifique su selección")
                                return
            
            columna["TIPO_DATO_GAMS"]=_tipoDato
            columna["NOMBRE_E_S"]=_nombre
            columna["TIPO_CAMPO"]=_tipoCampo
            columna["NOMBRE_COLUMNA"]=_nombreCol
            columna["FORMULA"]=_formula
            temCol.append(columna)
            
            self.ncol=self.ncol+1
            col=self.nombre+str(self.ncol)
            self.sabana[col]=temCol
            
            self.PrintColumnas()
            
            self.statusBar.showMessage("Nueva columna creada!")
        else:
            self.statusBar.showMessage("Nueva columna cancelada")

#%% main
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainProtoV2 = QtWidgets.QMainWindow()
    ui = Ui_MainProtoV2()
    ui.setupUi(MainProtoV2)
    ui.actionCerrar.triggered.connect(MainProtoV2.close)
    MainProtoV2.show()
    sys.exit(app.exec_())

